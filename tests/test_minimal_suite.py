from __future__ import annotations

import numpy as np
import pandas as pd
import pandas.testing as pdt

from v241_refactor_app.engine import _apply_withdrawal_trade, maybe_rebalance, run_historical_simulation, simulate_portfolio
from v241_refactor_app.mc_kernel import build_mc_path_plan, simulate_portfolio_path
from v241_refactor_app.models import HistoricalSelection
from v241_refactor_app.monte_carlo import simulate_monte_carlo, simulate_monte_carlo_summary

from conftest import make_assets, make_portfolio_inputs


def _first_positive_month(series: pd.Series) -> int | None:
    positive = series.gt(1e-9)
    if not positive.any():
        return None
    return int(np.flatnonzero(positive.to_numpy())[0] + 1)


def _first_zero_or_lower_month(series: pd.Series) -> int | None:
    non_positive = series.le(0.0)
    if not non_positive.any():
        return None
    return int(np.flatnonzero(non_positive.to_numpy())[0] + 1)


def _year_end_rows(selection: HistoricalSelection) -> pd.DataFrame:
    return selection.results_df.groupby("Year", sort=True).tail(1).reset_index(drop=True)


def test_monte_carlo_seed_regression_is_deterministic(base_inputs, base_assets, base_dataset):
    selection = run_historical_simulation(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )

    first = simulate_monte_carlo(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
    )
    second = simulate_monte_carlo(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
    )

    for left, right in zip(first, second):
        pdt.assert_frame_equal(left, right, check_dtype=False, check_exact=False, atol=1e-12, rtol=0.0)


def test_engine_and_kernel_are_invariant_for_same_path(base_inputs, base_assets, base_dataset):
    selection = run_historical_simulation(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )
    plan = build_mc_path_plan(base_inputs, base_assets, selection.filtered_periods[0])
    result = simulate_portfolio_path(
        plan=plan,
        portfolio_inputs=base_inputs,
        returns_matrix=selection.selected_returns_df.to_numpy(dtype=float),
        dividends_matrix=selection.selected_divs_df.to_numpy(dtype=float),
    )

    year_end = _year_end_rows(selection)
    assert np.allclose(year_end["Balance (USD)"].to_numpy(dtype=float), result.yearly_balances, atol=1e-9)
    assert np.allclose(year_end["Real Balance (USD)"].to_numpy(dtype=float), result.yearly_real_balances, atol=1e-9)

    expected_shortfall_month = _first_positive_month(selection.results_df["Withdrawal Shortfall (USD)"])
    expected_depletion_month = _first_zero_or_lower_month(selection.results_df["Balance (USD)"])
    expected_failure_month = min(
        [month for month in [expected_shortfall_month, expected_depletion_month] if month is not None],
        default=None,
    )

    assert result.shortfall_month == expected_shortfall_month
    assert result.depletion_month == expected_depletion_month
    assert result.failure_month == expected_failure_month


def test_threshold_band_does_not_rebalance_just_inside_band(base_inputs):
    balances = {"AAA": 54.999, "BBB": 45.001}
    target_weights = {"AAA": 50.0, "BBB": 50.0}
    asset_type_map = {"AAA": "Stock", "BBB": "Bond"}

    updated, ending_total, did_rebalance, cost_usd, turnover_pct = maybe_rebalance(
        balances=balances,
        target_weights=target_weights,
        method="Threshold Band",
        band=5.0,
        stage="month_end",
        rebalance_cost_bps=base_inputs.rebalance_cost_bps,
        asset_type_map=asset_type_map,
        portfolio_inputs=base_inputs,
    )

    assert updated == balances
    assert ending_total == 100.0
    assert did_rebalance is False
    assert cost_usd == 0.0
    assert turnover_pct == 0.0


def test_zero_balance_rebalance_is_safe_noop(base_inputs):
    updated, ending_total, did_rebalance, cost_usd, turnover_pct = maybe_rebalance(
        balances={"AAA": 0.0, "BBB": 0.0},
        target_weights={"AAA": 50.0, "BBB": 50.0},
        method="Annual",
        band=5.0,
        stage="year_end",
        rebalance_cost_bps=base_inputs.rebalance_cost_bps,
        asset_type_map={"AAA": "Stock", "BBB": "Bond"},
        portfolio_inputs=base_inputs,
    )

    assert updated == {"AAA": 0.0, "BBB": 0.0}
    assert ending_total == 0.0
    assert did_rebalance is False
    assert cost_usd == 0.0
    assert turnover_pct == 0.0


def test_withdrawal_shortfall_and_zero_balance_paths_never_go_negative():
    stressed_inputs = make_portfolio_inputs(
        initial_investment=1200.0,
        contribution_amount=0.0,
        withdrawal_mode="Fixed Dollar",
        withdrawal_rate=120.0,
        annual_fee_rate=0.0,
        tax_rate_dividends=0.0,
        tax_rate_withdrawals=0.0,
        cashflow_trade_cost_bps=0.0,
        rebalance_cost_bps=0.0,
        rebalancing_method="None",
        monte_carlo_years=1,
    )
    one_asset = make_assets((("AAA", 100.0, "Stock"),))
    plan = build_mc_path_plan(stressed_inputs, one_asset, pd.Timestamp("2020-01-31"))
    result = simulate_portfolio_path(
        plan=plan,
        portfolio_inputs=stressed_inputs,
        returns_matrix=np.zeros((12, 1), dtype=float),
        dividends_matrix=np.zeros((12, 1), dtype=float),
    )

    assert result.ending_balance == 0.0
    assert result.min_balance >= 0.0
    assert result.min_real_balance >= 0.0
    assert result.depletion_month == 10
    assert result.failure_month == 10

    remaining, gross_withdrawal, trade_cost, shortfall = _apply_withdrawal_trade(
        balances={"AAA": 100.0},
        desired_gross_amount=100.0,
        target_weights={"AAA": 100.0},
        asset_type_map={"AAA": "Stock"},
        base_bps=1000.0,
        portfolio_inputs=make_portfolio_inputs(cashflow_trade_cost_bps=1000.0),
    )
    assert gross_withdrawal < 100.0
    assert trade_cost > 0.0
    assert shortfall > 0.0
    assert remaining["AAA"] >= 0.0


def test_stock_data_filters_corporate_action_like_dividend_contamination(monkeypatch):
    import sys
    import types

    fake_streamlit = types.ModuleType("streamlit")
    def _cache_data(**_kwargs):
        def decorator(func):
            func.clear = lambda: None
            return func
        return decorator
    fake_streamlit.cache_data = _cache_data
    sys.modules.setdefault("streamlit", fake_streamlit)

    from v241_refactor_app import data_layer

    sample = [
        {"date": "2020-01-31T00:00:00.000Z", "adjClose": 100.0, "close": 100.0, "divCash": 0.0, "splitFactor": 1.0},
        {"date": "2020-02-28T00:00:00.000Z", "adjClose": 110.0, "close": 110.0, "divCash": 0.0, "splitFactor": 1.0},
        {"date": "2020-03-31T00:00:00.000Z", "adjClose": 121.0, "close": 60.5, "divCash": 2000.0, "splitFactor": 2.0},
        {"date": "2020-04-30T00:00:00.000Z", "adjClose": 133.1, "close": 66.55, "divCash": 0.0, "splitFactor": 1.0},
    ]

    monkeypatch.setattr(data_layer, "fetch_with_retry", lambda *args, **kwargs: sample)
    data_layer.get_stock_data.clear()

    price_returns, dividend_yields, years = data_layer.get_stock_data("GOOGL", "token")

    assert years == [2020]
    assert len(price_returns) == 3
    assert float(dividend_yields.max()) == 0.0
    assert float(price_returns.iloc[1]) > 0.0
    assert float(price_returns.iloc[1]) < 20.0


def test_parametric_var_95_is_reported_as_loss_sign():
    from v241_refactor_app.analytics import compute_risk_metrics

    periods = pd.date_range("2020-01-31", periods=6, freq="ME")
    results_df = pd.DataFrame(
        {
            "Period": periods,
            "Balance (USD)": [100, 105, 95, 110, 108, 115],
            "Portfolio Total Return (%)": [5.0, -9.5238095238, 15.7894736842, -1.8181818182, 6.4814814815, 0.0],
        }
    )

    metrics = compute_risk_metrics(results_df)
    assert metrics["Parametric VaR 95"] <= 0.0


def test_forward_overlay_custom_mode_reduces_drift_and_dividends():
    from v241_refactor_app.monte_carlo import _apply_forward_overlay, _effective_forward_spec

    assets = make_assets((("AAA", 60.0, "Stock"), ("BBB", 40.0, "Crypto")))
    historical_returns_df = pd.DataFrame(
        {
            "AAA": [2.0, 1.5, 2.5, 1.8, 2.2, 1.9],
            "BBB": [5.0, 4.5, 6.0, 5.5, 4.8, 5.2],
        }
    )
    sampled_returns = historical_returns_df.to_numpy(dtype=float)
    sampled_dividends = np.array([[0.20, 0.0]] * len(historical_returns_df), dtype=float)
    inputs = make_portfolio_inputs(
        monte_carlo_forward_mode="Custom Forward Stress",
        monte_carlo_return_haircut_pct=50.0,
        monte_carlo_return_shift_pct=-3.0,
        monte_carlo_vol_multiplier=1.25,
        monte_carlo_growth_extra_haircut_pct=10.0,
        monte_carlo_crypto_extra_haircut_pct=15.0,
        monte_carlo_dividend_multiplier=0.5,
    )

    forward_spec = _effective_forward_spec(inputs)
    adjusted_returns, adjusted_dividends = _apply_forward_overlay(
        sampled_returns=sampled_returns,
        sampled_dividends=sampled_dividends,
        historical_returns_df=historical_returns_df,
        assets=assets,
        forward_spec=forward_spec,
    )

    assert adjusted_returns.shape == sampled_returns.shape
    assert adjusted_dividends.shape == sampled_dividends.shape
    assert float(adjusted_returns[:, 0].mean()) < float(sampled_returns[:, 0].mean())
    assert float(adjusted_returns[:, 1].mean()) < float(sampled_returns[:, 1].mean())
    assert float(adjusted_dividends[:, 0].max()) == 0.10
    assert float(adjusted_dividends[:, 1].max()) == 0.0


def test_forward_stress_mode_lowers_mc_median_path(base_assets, base_dataset):
    historical_inputs = make_portfolio_inputs(
        contribution_amount=0.0,
        withdrawal_rate=0.0,
        annual_fee_rate=0.0,
        tax_rate_dividends=0.0,
        tax_rate_withdrawals=0.0,
        rebalance_cost_bps=0.0,
        cashflow_trade_cost_bps=0.0,
        rebalancing_method="Annual",
        monte_carlo_sims=64,
        monte_carlo_years=2,
        monte_carlo_seed=42,
        monte_carlo_adaptive_convergence=False,
        monte_carlo_forward_mode="Historical Base",
    )
    stressed_inputs = make_portfolio_inputs(
        contribution_amount=0.0,
        withdrawal_rate=0.0,
        annual_fee_rate=0.0,
        tax_rate_dividends=0.0,
        tax_rate_withdrawals=0.0,
        rebalance_cost_bps=0.0,
        cashflow_trade_cost_bps=0.0,
        rebalancing_method="Annual",
        monte_carlo_sims=64,
        monte_carlo_years=2,
        monte_carlo_seed=42,
        monte_carlo_adaptive_convergence=False,
        monte_carlo_forward_mode="Stagnation & De-Rating",
    )

    selection = run_historical_simulation(
        portfolio_inputs=historical_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )

    base_percentiles, _, _, _ = simulate_monte_carlo(
        portfolio_inputs=historical_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
    )
    stressed_percentiles, _, _, _ = simulate_monte_carlo(
        portfolio_inputs=stressed_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
    )

    assert float(stressed_percentiles["Median"].iloc[-1]) < float(base_percentiles["Median"].iloc[-1])


def test_bucket_cma_targets_forward_mode_applies_bucket_specific_targets():
    from v241_refactor_app.monte_carlo import _apply_forward_overlay, _effective_forward_spec

    assets = make_assets((('AAA', 50.0, 'Stock'), ('BBB', 50.0, 'Crypto')))
    historical_returns_df = pd.DataFrame(
        {
            'AAA': [0.8, 1.0, 1.2, 0.9, 1.1, 1.0],
            'BBB': [3.0, 4.0, 5.0, 4.5, 3.5, 4.2],
        }
    )
    sampled_returns = historical_returns_df.to_numpy(dtype=float)
    sampled_dividends = np.array([[0.20, 0.0]] * len(historical_returns_df), dtype=float)
    inputs = make_portfolio_inputs(
        monte_carlo_forward_mode='Bucket CMA Targets',
        monte_carlo_bucket_cma_core_return_pct=6.0,
        monte_carlo_bucket_cma_crypto_return_pct=0.0,
        monte_carlo_bucket_cma_core_vol_multiplier=1.0,
        monte_carlo_bucket_cma_crypto_vol_multiplier=1.0,
        monte_carlo_dividend_multiplier=0.5,
    )

    forward_spec = _effective_forward_spec(inputs)
    adjusted_returns, adjusted_dividends = _apply_forward_overlay(
        sampled_returns=sampled_returns,
        sampled_dividends=sampled_dividends,
        historical_returns_df=historical_returns_df,
        assets=assets,
        forward_spec=forward_spec,
    )

    aaa_target_monthly = (((1.0 + 0.06) ** (1.0 / 12.0)) - 1.0) * 100.0
    bbb_target_monthly = 0.0
    assert abs(float(adjusted_returns[:, 0].mean()) - aaa_target_monthly) < 1e-9
    assert abs(float(adjusted_returns[:, 1].mean()) - bbb_target_monthly) < 1e-9
    assert float(adjusted_dividends[:, 0].max()) == 0.10



def test_fragility_analysis_returns_ranked_scenarios(base_inputs, base_assets, base_dataset):
    from v241_refactor_app.analysis_lab import build_fragility_analysis

    selection = run_historical_simulation(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )
    outputs = build_fragility_analysis(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
    )

    fragility_df = outputs['fragility_df']
    fragility_pivot_df = outputs['fragility_pivot_df']
    assert not fragility_df.empty
    assert fragility_df.iloc[0]['Scenario'] == 'Base Case'
    assert 'Fragility Rank' in fragility_df.columns
    assert not fragility_pivot_df.empty



def test_decision_engine_ranks_policies(base_inputs, base_assets, base_dataset):
    from v241_refactor_app.analysis_lab import build_decision_policy_analysis

    selection = run_historical_simulation(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )
    outputs = build_decision_policy_analysis(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
        objective='Balanced robustness',
    )

    policy_df = outputs['policy_df']
    recommendation_df = outputs['recommendation_df']
    assert not policy_df.empty
    assert int(policy_df.iloc[0]['Rank']) == 1
    assert 'Current' in set(policy_df['Policy'])
    assert not recommendation_df.empty


def test_app_input_builder_does_not_zero_out_inflation_when_real_display_is_off():
    from v241_refactor_app.app import _build_portfolio_inputs

    inputs = _build_portfolio_inputs(
        initial_investment=100000.0,
        withdrawal_rate=4.0,
        reinvest_dividends=False,
        contribution_amount=0.0,
        contribution_end_year_input="",
        annual_fee_rate=0.0,
        inflation_rate=3.5,
        tax_rate_dividends=0.0,
        tax_rate_withdrawals=0.0,
        withdrawal_mode="Fixed Dollar",
        show_benchmark_overlay=False,
        benchmark_ticker="SPY",
        benchmark_type="Stock",
        rebalancing_method="Annual",
        rebalance_band=5.0,
        rebalance_cost_bps=5.0,
        cashflow_trade_cost_bps=0.0,
        asset_class_aware_costs=True,
        aum_cost_scaling=True,
        aum_cost_scaling_strength=0.25,
        crypto_cost_multiplier=4.0,
        analysis_mode="Historical Backtest",
        monte_carlo_sims=100,
        monte_carlo_years=30,
        monte_carlo_seed=42,
        monte_carlo_block_size_months=12,
        monte_carlo_regime_mode="All History",
        monte_carlo_regime_window_months=6,
        monte_carlo_bootstrap_method="Block Bootstrap",
        monte_carlo_regime_strength=1.0,
        monte_carlo_adaptive_convergence=True,
        monte_carlo_target_stderr_pct=0.35,
        monte_carlo_forward_mode="Historical Base",
        monte_carlo_return_haircut_pct=0.0,
        monte_carlo_return_shift_pct=0.0,
        monte_carlo_vol_multiplier=1.0,
        monte_carlo_growth_extra_haircut_pct=0.0,
        monte_carlo_crypto_extra_haircut_pct=0.0,
        monte_carlo_dividend_multiplier=1.0,
        monte_carlo_bucket_cma_defensive_return_pct=4.0,
        monte_carlo_bucket_cma_core_return_pct=6.0,
        monte_carlo_bucket_cma_aggressive_return_pct=8.0,
        monte_carlo_bucket_cma_crypto_return_pct=12.0,
        monte_carlo_bucket_cma_defensive_vol_multiplier=0.9,
        monte_carlo_bucket_cma_core_vol_multiplier=1.0,
        monte_carlo_bucket_cma_aggressive_vol_multiplier=1.1,
        monte_carlo_bucket_cma_crypto_vol_multiplier=1.25,
    )
    assert inputs.inflation_rate == 3.5


def test_withdrawal_tax_targets_net_cash_not_gross_sales():
    taxed_inputs = make_portfolio_inputs(
        initial_investment=120000.0,
        contribution_amount=0.0,
        withdrawal_mode="Fixed Dollar",
        withdrawal_rate=12.0,
        annual_fee_rate=0.0,
        tax_rate_dividends=0.0,
        tax_rate_withdrawals=20.0,
        cashflow_trade_cost_bps=0.0,
        rebalance_cost_bps=0.0,
        rebalancing_method="None",
        monte_carlo_years=1,
    )
    one_asset = make_assets((("AAA", 100.0, "Stock"),))
    plan = build_mc_path_plan(taxed_inputs, one_asset, pd.Timestamp("2020-01-31"))
    result = simulate_portfolio_path(
        plan=plan,
        portfolio_inputs=taxed_inputs,
        returns_matrix=np.zeros((12, 1), dtype=float),
        dividends_matrix=np.zeros((12, 1), dtype=float),
    )
    assert result.shortfall == 0.0

    periods = list(pd.date_range("2020-01-31", periods=12, freq="ME"))
    results_df = simulate_portfolio(
        portfolio_inputs=taxed_inputs,
        assets=one_asset,
        returns_df=pd.DataFrame({"AAA": np.zeros(12)}, index=periods),
        dividends_df=pd.DataFrame({"AAA": np.zeros(12)}, index=periods),
        filtered_periods=periods,
    )
    first_row = results_df.iloc[0]
    assert abs(float(first_row["Withdrawal Target (USD)"]) - 1200.0) < 1e-9
    assert abs(float(first_row["Withdrawal (USD)"]) - 1200.0) < 1e-9
    assert abs(float(first_row["Gross Withdrawal (USD)"]) - 1500.0) < 1e-9
    assert abs(float(first_row["Withdrawal Shortfall (USD)"])) < 1e-9


def test_fragility_grid_explicit_shift_axis_is_present(base_inputs, base_assets, base_dataset):
    from v241_refactor_app.analysis_lab import build_fragility_analysis

    selection = run_historical_simulation(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )
    outputs = build_fragility_analysis(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
    )
    assert "Explicit Forward Return Shift (%)" in outputs["fragility_grid_df"].columns


def test_asset_matrix_diagnostics_surface_fallback_and_dividend_filters(monkeypatch):
    from v241_refactor_app.data_layer import build_asset_matrices

    stock_index = pd.date_range("2020-01-31", periods=12, freq="ME")
    stock_returns = pd.Series(np.linspace(0.5, 1.6, 12), index=stock_index)
    stock_divs = pd.Series([0.0] * 11 + [0.2], index=stock_index)

    def fake_stock_meta(ticker, token):
        return stock_returns, stock_divs, [2020], {
            "data_source": "stock_api",
            "filtered_dividend_events": 2,
            "clipped_dividend_months": 1,
            "fallback_used": False,
            "fallback_reason": "",
        }

    def fake_crypto_meta(ticker, token):
        return stock_returns, [2020], {
            "data_source": "stock_api_fallback",
            "filtered_dividend_events": 0,
            "clipped_dividend_months": 0,
            "fallback_used": True,
            "fallback_reason": "crypto endpoint unavailable",
        }

    monkeypatch.setattr("v241_refactor_app.data_layer._get_stock_data_with_metadata", fake_stock_meta)
    monkeypatch.setattr("v241_refactor_app.data_layer._get_crypto_data_with_metadata", fake_crypto_meta)

    _, _, _, _, diagnostics = build_asset_matrices(
        [
            make_assets((("AAA", 50.0, "Stock"),))[0],
            make_assets((("BTC", 50.0, "Crypto"),))[0],
        ],
        token="dummy",
    )
    diagnostics_df = pd.DataFrame(diagnostics)
    assert "Data Source" in diagnostics_df.columns
    assert bool(diagnostics_df.loc[diagnostics_df["Ticker"] == "BTC", "Fallback Used"].iloc[0]) is True
    assert int(diagnostics_df.loc[diagnostics_df["Ticker"] == "AAA", "Filtered Dividend Events"].iloc[0]) == 2



def test_export_workbook_includes_decision_lab_and_provenance_sheets(base_inputs, base_assets, base_dataset):
    import io
    from openpyxl import load_workbook

    from v241_refactor_app.analysis_lab import build_decision_policy_analysis, build_fragility_analysis
    from v241_refactor_app.analytics import compute_risk_metrics, compute_summary_metrics
    from v241_refactor_app.orchestration import build_export_bytes_cached
    from v241_refactor_app.utils import serialize_assets, serialize_portfolio_inputs

    selection = run_historical_simulation(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )
    fragility_outputs = build_fragility_analysis(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
    )
    decision_outputs = build_decision_policy_analysis(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
        objective="Balanced robustness",
    )
    metrics = compute_summary_metrics(selection.results_df, selection.filtered_periods)
    risk_metrics = compute_risk_metrics(selection.results_df)
    risk_table = pd.DataFrame({"Metric": list(risk_metrics.keys()), "Value": list(risk_metrics.values())})
    diagnostics_df = pd.DataFrame([
        {
            "Ticker": "AAA",
            "Data Source": "stock_api",
            "Filtered Dividend Events": 0,
            "Clipped Dividend Months": 0,
            "Fallback Used": False,
            "Fallback Reason": "",
            "Overlap Start": "2020-01",
            "Overlap End": "2021-12",
            "Overlap Months": 24,
        }
    ])

    workbook_bytes = build_export_bytes_cached(
        portfolio_input_dict=serialize_portfolio_inputs(base_inputs),
        asset_specs=serialize_assets(base_assets),
        year_range=(2020, 2021),
        results_df=selection.results_df,
        component_df=selection.component_df,
        selected_returns_df=selection.selected_returns_df,
        selected_divs_df=selection.selected_divs_df,
        weighted_returns=selection.weighted_returns,
        weighted_divs=selection.weighted_divs,
        metrics=metrics,
        diagnostics_df=diagnostics_df,
        risk_table=risk_table,
        fragility_df=fragility_outputs["fragility_df"],
        fragility_pivot_df=fragility_outputs["fragility_pivot_df"],
        policy_df=decision_outputs["policy_df"],
        recommendation_df=decision_outputs["recommendation_df"],
        decision_objective="Balanced robustness",
    )

    wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True)
    sheet_names = set(wb.sheetnames)
    expected = {
        "Run Provenance",
        "Forward Audit",
        "Decision Lab Settings",
        "Fragility Summary",
        "Fragility Matrix",
        "Policy Comparison",
        "Policy Recommendation",
    }
    assert expected.issubset(sheet_names)



def test_export_signature_changes_when_decision_lab_state_changes():
    from v241_refactor_app.app import _export_signature

    base = _export_signature(
        active_core_signature="core",
        year_range=(2020, 2021),
        benchmark_signature=None,
        mc_signature=None,
        scenario_signature=None,
        fragility_signature=None,
        decision_signature=None,
    )
    with_fragility = _export_signature(
        active_core_signature="core",
        year_range=(2020, 2021),
        benchmark_signature=None,
        mc_signature=None,
        scenario_signature=None,
        fragility_signature="frag-a",
        decision_signature=None,
    )
    with_decision = _export_signature(
        active_core_signature="core",
        year_range=(2020, 2021),
        benchmark_signature=None,
        mc_signature=None,
        scenario_signature=None,
        fragility_signature="frag-a",
        decision_signature="decision-b",
    )

    assert base != with_fragility
    assert with_fragility != with_decision



def test_summary_only_monte_carlo_matches_full_summary(base_inputs, base_assets, base_dataset):
    selection = run_historical_simulation(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )
    _, full_summary_df, _, _ = simulate_monte_carlo(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
    )
    summary_only_df = simulate_monte_carlo_summary(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
    )

    merged = full_summary_df.merge(summary_only_df, on=["Metric", "Unit"], suffixes=("_full", "_summary"))
    compare_metrics = [
        "Median Ending Balance",
        "Real Median Ending Balance",
        "10th Percentile Ending Balance",
        "Real 10th Percentile Ending Balance",
        "90th Percentile Ending Balance",
        "Real 90th Percentile Ending Balance",
        "Failure Rate (Ruin or Shortfall)",
        "Failure StdErr",
        "Ruin Rate",
        "Spending Shortfall Rate",
        "Median Portfolio CAGR",
        "CVaR 5% Ending Balance",
        "Median Minimum Balance",
        "Median Minimum Real Balance",
        "Median Failure Year (conditional)",
        "Median Depletion Year (conditional)",
        "Median Shortfall Year (conditional)",
    ]
    merged = merged[merged["Metric"].isin(compare_metrics)].copy()
    assert not merged.empty
    assert np.allclose(merged["Value_full"].to_numpy(dtype=float), merged["Value_summary"].to_numpy(dtype=float), atol=1e-12, rtol=0.0, equal_nan=True)



def test_fragility_fast_mode_uses_smaller_surface(base_inputs, base_assets, base_dataset):
    from v241_refactor_app.analysis_lab import build_fragility_analysis

    selection = run_historical_simulation(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )
    full_outputs = build_fragility_analysis(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
        analysis_mode="Full",
    )
    fast_outputs = build_fragility_analysis(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
        analysis_mode="Fast",
    )

    assert fast_outputs.get("fragility_mode") == "Fast"
    assert full_outputs.get("fragility_mode") == "Full"
    assert len(fast_outputs["fragility_df"]) < len(full_outputs["fragility_df"])
    assert len(fast_outputs["fragility_grid_df"]) < len(full_outputs["fragility_grid_df"])
    assert tuple(fast_outputs["fragility_pivot_df"].shape) == (3, 3)
    assert tuple(full_outputs["fragility_pivot_df"].shape) == (5, 5)
