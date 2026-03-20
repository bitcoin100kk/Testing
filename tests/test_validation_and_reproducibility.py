from __future__ import annotations

import io
import tomllib
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from v241_refactor_app.analysis_lab import build_decision_policy_analysis
from v241_refactor_app.engine import run_historical_simulation
from v241_refactor_app.monte_carlo import build_monte_carlo_validation_report, simulate_monte_carlo
from v241_refactor_app.orchestration import build_export_bytes_cached, simulate_monte_carlo_cached
from v241_refactor_app.utils import serialize_assets, serialize_portfolio_inputs
from v241_refactor_app.version import __version__


def test_monte_carlo_validation_report_contains_confidence_intervals(base_inputs, base_assets, base_dataset):
    selection = run_historical_simulation(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )
    percentiles_df, summary_df, paths_df, convergence_df = simulate_monte_carlo(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
    )
    validation_df = build_monte_carlo_validation_report(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        summary_df=summary_df,
        paths_df=paths_df,
        convergence_df=convergence_df,
    )

    assert not validation_df.empty
    metrics = set(validation_df["Metric"])
    assert "Failure rate 95% CI lower" in metrics
    assert "Failure rate 95% CI upper" in metrics
    assert "Effective distinct start months" in metrics
    completed = float(validation_df.loc[validation_df["Metric"].eq("Completed simulations"), "Value"].iloc[0])
    assert completed > 0


def test_cached_monte_carlo_artifacts_include_validation_df(base_inputs, base_assets, base_dataset):
    selection = run_historical_simulation(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )
    outputs = simulate_monte_carlo_cached(
        portfolio_input_dict=serialize_portfolio_inputs(base_inputs),
        asset_specs=serialize_assets(base_assets),
        selected_returns_df=selection.selected_returns_df,
        selected_divs_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
    )
    assert "validation_df" in outputs
    assert not outputs["validation_df"].empty


def test_decision_policy_analysis_includes_explicit_objective_metadata(base_inputs, base_assets, base_dataset):
    selection = run_historical_simulation(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )
    outputs = build_decision_policy_analysis(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
        objective="Balanced robustness",
    )
    policy_df = outputs["policy_df"]
    recommendation_df = outputs["recommendation_df"]

    assert {"Objective", "Objective Rank", "Pareto Efficient", "Recommendation Basis"}.issubset(policy_df.columns)
    assert bool(policy_df["Pareto Efficient"].any()) is True
    assert bool(recommendation_df["Objective Description"].astype(str).str.len().gt(0).all()) is True


def test_export_workbook_includes_mc_validation_and_version(base_inputs, base_assets, base_dataset):
    selection = run_historical_simulation(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        dataset=base_dataset,
        selected_range=(2020, 2021),
    )
    percentiles_df, summary_df, paths_df, convergence_df = simulate_monte_carlo(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        start_period=selection.filtered_periods[0],
    )
    validation_df = build_monte_carlo_validation_report(
        portfolio_inputs=base_inputs,
        assets=base_assets,
        historical_returns_df=selection.selected_returns_df,
        historical_dividends_df=selection.selected_divs_df,
        summary_df=summary_df,
        paths_df=paths_df,
        convergence_df=convergence_df,
    )

    workbook_bytes = build_export_bytes_cached(
        portfolio_input_dict=serialize_portfolio_inputs(base_inputs),
        asset_specs=serialize_assets(base_assets),
        year_range=(2020, 2021),
        results_df=selection.results_df,
        component_df=selection.component_df,
        selected_returns_df=selection.selected_returns_df,
        selected_divs_df=selection.selected_divs_df,
        weighted_returns=selection.weighted_returns,
        weighted_divs=selection.weighted_divs,
        metrics={"Final Balance": 1.0},
        mc_percentiles_df=percentiles_df,
        mc_summary_df=summary_df,
        mc_paths_df=paths_df,
        mc_convergence_df=convergence_df,
        mc_validation_df=validation_df,
    )

    wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True)
    assert "Monte Carlo Validation" in wb.sheetnames
    ws = wb["Run Provenance"]
    values = list(ws.iter_rows(values_only=True))
    flattened = [item for row in values for item in row if item is not None]
    assert __version__ in flattened


def test_pyproject_version_matches_package_version():
    pyproject = tomllib.loads(Path("pyproject.toml").read_text())
    assert pyproject["project"]["version"] == __version__
